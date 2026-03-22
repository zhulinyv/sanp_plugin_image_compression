import os
import shutil
from pathlib import Path

import cv2
import numpy as np
import openpyxl
from openpyxl.drawing.image import Image as OPXIMG
from openpyxl.styles import Alignment
from PIL import Image as PILIMG

from utils.imgtools import return_pnginfo, revert_img_info
from utils.prepare import logger
from utils.utils import file_namel2pathl, file_path2list


def _image_compression(format_, image):
    # 读取图像，保留所有通道（包括 Alpha）
    cv2_image = cv2.imread(image, cv2.IMREAD_UNCHANGED)

    if format_ == "jpg":
        # JPG 不支持透明，如果有 Alpha 通道，先与白色背景合并
        if cv2_image.shape[2] == 4:
            # 创建白色背景
            bg = np.ones(cv2_image.shape[:2] + (3,), dtype=np.uint8) * 255
            # 分离 BGRA 通道
            b, g, r, a = cv2.split(cv2_image)
            # 将 Alpha 归一化到 0-1 并合成
            alpha = a / 255.0
            b = (b * alpha + 255 * (1 - alpha)).astype(np.uint8)
            g = (g * alpha + 255 * (1 - alpha)).astype(np.uint8)
            r = (r * alpha + 255 * (1 - alpha)).astype(np.uint8)
            cv2_image = cv2.merge([b, g, r])
        else:
            cv2_image = cv2.cvtColor(cv2_image, cv2.COLOR_BGRA2BGR) if cv2_image.shape[2] == 4 else cv2_image

        compression_params = [cv2.IMWRITE_JPEG_QUALITY, 90]
    elif format_ == "png":
        # PNG 直接保留 Alpha 通道
        compression_params = [cv2.IMWRITE_PNG_COMPRESSION, 9]
    else:
        raise ValueError("Unsupported format")

    # 确保输出目录存在
    os.makedirs("./output", exist_ok=True)
    temp_path = f"./output/temp.{format_}"
    cv2.imwrite(temp_path, cv2_image, compression_params)

    # 保留原 PNG 元数据（如果需要）
    if format_ == "png":
        with PILIMG.open(image) as pil_img:
            pnginfo = pil_img.info
            revert_img_info(None, temp_path, pnginfo)  # 假设 revert_img_info 函数存在

    # 替换原文件
    os.remove(image)
    base, _ = os.path.splitext(image)
    shutil.move(temp_path, base + f".{format_}")


def image_compression(format_, image_path):
    image_list = file_namel2pathl(file_path2list(image_path), image_path)
    for image in image_list:
        logger.info(f"正在压缩 {image} ...")
        _image_compression(format_, image)
    logger.success("压缩完成!")
    return "压缩完成!"


def image_organization(format_, image_path, switch):
    logger.info("正在整理图片...")
    image_list = file_namel2pathl(file_path2list(image_path), image_path)
    iamge_number = len(image_list)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.row_dimensions[1].height = 50
    for row in range(2, iamge_number + 2):
        ws.row_dimensions[row].height = 300
    for col in ["A", "B", "C"]:
        ws.column_dimensions[col].width = 40
    for col in range(4, 14):
        col_letter = openpyxl.utils.get_column_letter(col)
        ws.column_dimensions[col_letter].width = 20
    ws.append(
        [
            "图片",
            "正面提示词",
            "负面提示词",
            "分辨率",
            "采样步数",
            "提示词相关性",
            "噪声计划表",
            "采样器",
            "sm",
            "sm_dyn",
            "variety",
            "decrisp",
            "随机种子",
        ]
    )
    number = 2
    for image in image_list:
        with PILIMG.open(image) as pilimg:
            (
                positive_input,
                negative_input,
                width,
                height,
                steps,
                scale,
                noise_schedule,
                sampler,
                sm,
                sm_dyn,
                variety,
                decrisp,
                seed,
            ) = return_pnginfo(pilimg)[:-1]
            w, h = pilimg.size
        if switch:
            _image_compression(format_, image)
            image = OPXIMG(str(image)[:-4] + f".{format_}")
        else:
            image = OPXIMG(image)
        image.width, image.height = 260, int(260 / w * h)
        ws.add_image(image, f"A{number}")
        ws[f"B{number}"] = positive_input
        ws[f"C{number}"] = negative_input
        ws[f"D{number}"] = f"{width}x{height}"
        ws[f"E{number}"] = steps
        ws[f"F{number}"] = scale
        ws[f"G{number}"] = noise_schedule
        ws[f"H{number}"] = sampler
        ws[f"I{number}"] = sm
        ws[f"J{number}"] = sm_dyn
        ws[f"K{number}"] = variety
        ws[f"L{number}"] = decrisp
        ws[f"M{number}"] = seed
        number += 1
    alignment = Alignment(horizontal="center", vertical="center", wrapText=True)
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = alignment
    wb.save(Path(image_path) / "organization.xlsx")
    logger.success("整理完成!")
    return "整理完成!"
